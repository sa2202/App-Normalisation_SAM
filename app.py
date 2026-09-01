"""
SAM Application Normalization Accelerator
==========================================
Run:  streamlit run app.py
"""
import io
import os
import tempfile
import warnings
import pandas as pd
import streamlit as st

warnings.filterwarnings("ignore")

from normalizer import ProductNormalizer, BACKEND
from schema_detect import sniff_and_load, detect_schema
from inventory_import import classify_inventory
from file_evidence import process_file_evidence, looks_like_file_evidence
from bulk_import import preview_import, commit_import

# ── constants ─────────────────────────────────────────────────────────────
LIBRARY_PATH = "canonical_library.csv"
APP_VERSION  = "1.0"

# ── session state defaults ────────────────────────────────────────────────
for key in ["normalizer", "results", "source_df", "llm_results", "aliases_added_this_session",
            "itam_preview", "fe_out"]:
    if key not in st.session_state:
        st.session_state[key] = None

if st.session_state.normalizer is None:
    st.session_state.normalizer = ProductNormalizer(LIBRARY_PATH)

norm = st.session_state.normalizer

# ── page config ───────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SAM Normalization Accelerator",
    page_icon="🔎",
    layout="wide",
)

# ── sidebar ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🔎 SAM Normalizer")
    st.caption(f"v{APP_VERSION}")
    st.divider()

    # Library stats
    pub_counts = norm.lib["publisher"].value_counts()
    main_pubs  = pub_counts[~pub_counts.index.str.startswith("(")]
    st.metric("Products in library", len(norm.lib))
    st.metric("Publishers covered", len(main_pubs))
    st.write(f"**Matching backend:** `{BACKEND}`")
    if "rapidfuzz" not in BACKEND:
        st.warning("Install `rapidfuzz` for better accuracy:\n```\npip install rapidfuzz\n```")

    with st.expander("View library"):
        st.dataframe(
            norm.lib[["publisher","product_family","edition","version","metric_type"]],
            use_container_width=True, height=300
        )

    # ── Library persistence ────────────────────────────────────────────
    # On Streamlit Cloud the filesystem is EPHEMERAL: anything written here
    # (confirmed aliases, bulk imports) is wiped whenever the container
    # rebuilds — on redeploy, on wake-from-sleep, on maintenance.
    # So changes must be downloaded and committed back to GitHub to persist.
    added = st.session_state.get("aliases_added_this_session", 0)
    if added:
        st.warning(
            f"⚠️ **{added} change(s) made this session — not yet permanent.**\n\n"
            "Download the library below and commit it to GitHub, or these "
            "will be lost when the app restarts."
        )
    st.download_button(
        f"⬇️ Download library{' ⚠️' if added else ''}",
        norm.lib.to_csv(index=False).encode(),
        "canonical_library.csv",
        "text/csv",
        use_container_width=True,
        type="primary" if added else "secondary",
        help="Download, then commit to GitHub to make changes permanent for the whole team.",
    )
    with st.expander("ℹ️ How to make changes permanent"):
        st.markdown(
            "Streamlit Cloud wipes file changes on every restart, so teaching "
            "the library only lasts for this session unless you sync it back.\n\n"
            "**After a teaching session:**\n"
            "1. Click **Download library** above\n"
            "2. Go to your GitHub repo → click `canonical_library.csv`\n"
            "3. Click the ✏️ pencil (edit) → delete all contents\n"
            "4. Open the downloaded file, copy everything, paste it in\n"
            "5. **Commit changes**\n\n"
            "The app redeploys automatically and everyone gets the updated library.\n\n"
            "*Running locally instead? Changes save automatically — no sync needed.*"
        )
    st.divider()

    # Bulk import from ITAM tool
    st.subheader("📥 Grow library from ITAM export")
    st.caption("Flexera / Snow / ServiceNow SAM / any normalized product list")
    itam_file = st.file_uploader("ITAM export (CSV/Excel)", type=["csv","xlsx","xls"], key="itam_up")
    if itam_file:
        itam_df = pd.read_csv(itam_file) if itam_file.name.endswith(".csv") else pd.read_excel(itam_file)
        st.write(f"{len(itam_df)} rows · {len(itam_df.columns)} columns")
        cols_none = ["(none)"] + list(itam_df.columns)
        def pick(hints):
            for h in hints:
                for c in itam_df.columns:
                    if h.lower() in str(c).lower(): return cols_none.index(c) + 1
            return 0
        pub_col    = st.selectbox("Publisher column",         cols_none, index=pick(["publisher","vendor","manufacturer"]))
        prod_col   = st.selectbox("Product / Title column",   cols_none, index=pick(["product","title","name","software"]))
        ed_col     = st.selectbox("Edition (optional)",       cols_none, index=pick(["edition","sku"]))
        ver_col    = st.selectbox("Version (optional)",       cols_none, index=pick(["version","ver","release"]))
        metric_col = st.selectbox("License metric (optional)", cols_none, index=pick(["metric","license metric"]))

        if st.button("Preview import"):
            col_map = {
                "publisher": pub_col    if pub_col    != "(none)" else None,
                "product":   prod_col   if prod_col   != "(none)" else None,
                "edition":   ed_col     if ed_col     != "(none)" else None,
                "version":   ver_col    if ver_col    != "(none)" else None,
                "metric":    metric_col if metric_col != "(none)" else None,
            }
            st.session_state.itam_preview = preview_import(itam_df, norm.lib, col_map)

        if st.session_state.itam_preview is not None:
            p = st.session_state.itam_preview
            st.write(f"**{len(p['new_rows'])}** new  ·  {p['duplicate_count']} already in library  ·  {p['blank_count']} skipped")
            st.dataframe(p["new_rows"], use_container_width=True, height=200)
            if len(p["new_rows"]) and st.button("Add to library ✅", type="primary"):
                added = commit_import(norm, p["new_rows"], LIBRARY_PATH)
                st.session_state.aliases_added_this_session = (
                    st.session_state.get("aliases_added_this_session", 0) + added
                )
                st.success(f"Added {added} products to the library.")
                st.session_state.itam_preview = None
                st.rerun()

# ── main ──────────────────────────────────────────────────────────────────
st.title("SAM Application Normalization Accelerator")
st.caption(
    "Upload a client's software inventory in **any format** — Flexera, "
    "ServiceNow, SCCM, raw script dumps, or file evidence — and get a "
    "normalized, confidence-scored product list ready for ELP work."
)

# ── Step 1: upload ─────────────────────────────────────────────────────────
st.subheader("1 · Upload client data")
mode = st.radio(
    "What kind of data is this?",
    options=[
        "📋  Software inventory  (Name / Vendor / Version — one row per installed product)",
        "📁  File evidence  (Name / Version / Path — one row per file, from Flexera FileEvidence or similar)",
        "🗂️  Generic CSV / Excel  (I will map columns manually)",
    ],
    label_visibility="collapsed",
)

uploaded = st.file_uploader(
    "Drop file here", type=["csv","xlsx","xls","txt","tsv"],
    label_visibility="collapsed"
)

# ── File evidence path ─────────────────────────────────────────────────────
if "File evidence" in mode:
    if uploaded:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded.name)[1]) as tmp:
            tmp.write(uploaded.getvalue()); tmp_path = tmp.name
        try:
            raw_df = sniff_and_load(tmp_path)
        finally:
            os.unlink(tmp_path)

        st.success(f"Loaded {len(raw_df)} rows · columns: {', '.join(raw_df.columns)}")
        if not looks_like_file_evidence(raw_df):
            st.warning("⚠️ This doesn't look like file evidence — expected a column of filenames (`.exe`/`.sys`/`.dll`) and a Path column. Check you picked the right mode.")

        cols = list(raw_df.columns)
        def _idx(hints, default=0):
            for h in hints:
                for i, c in enumerate(cols):
                    if h.lower() in str(c).lower(): return i
            return default
        c1, c2 = st.columns(2)
        with c1:
            fe_name = st.selectbox("File name column", cols, index=_idx(["name","file name"]))
            fe_ver  = st.selectbox("Version column",   cols, index=_idx(["version"], 1 if len(cols)>1 else 0))
        with c2:
            fe_path = st.selectbox("Path column",      cols, index=_idx(["path","location","directory"], 2 if len(cols)>2 else 0))
            fe_dev  = st.selectbox("Device / Host column (optional)", ["(none)"]+cols)

        if st.button("Process file evidence 🔍", type="primary"):
            with st.spinner("Filtering, identifying, deduplicating…"):
                st.session_state.fe_out = process_file_evidence(
                    raw_df,
                    name_col=fe_name, version_col=fe_ver, path_col=fe_path,
                    device_col=None if fe_dev=="(none)" else fe_dev,
                )

    if st.session_state.fe_out:
        fe = st.session_state.fe_out
        s  = fe["stats"]
        st.divider()
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Input rows",         s["input_rows"])
        m2.metric("Excluded (non-install)", s["excluded_paths"])
        m3.metric("Noise filenames",    s["noise_files"])
        m4.metric("Identified rows",    s["identified_rows"])
        m5.metric("Distinct products",  s["distinct_products"])

        st.info(
            f"**{s['identified_rows']} file rows → {s['distinct_products']} products** after deduplication.  "
            f"{s['excluded_paths']} rows were in locations that aren't real installs "
            f"(Recycle Bin, Windows.old, driver store, Downloads, backup snapshots).  "
            f"{s['noise_files']} had noise filenames (GUID-named, `$`-prefixed, setup/uninstall files). "
            f"Counting those would have inflated deployment counts."
        )

        st.subheader("Deduplicated deployment picture")
        st.dataframe(fe["products"], use_container_width=True)
        st.caption(
            "⚠️  *Supporting files only* = no product executable found — confirm the product "
            "is genuinely installed.  File versions ≠ product versions."
        )

        unid = fe["detail"][fe["detail"]["status"]=="UNIDENTIFIED"]
        if len(unid):
            with st.expander(f"🔍 {len(unid)} unidentified files — add recurring ones to FILE_SIGNATURES in file_evidence.py"):
                st.dataframe(unid[["file_name","file_version","path"]], use_container_width=True)

        with st.expander("Full audit trail (every input row annotated)"):
            st.dataframe(fe["detail"], use_container_width=True)

        b1, b2 = st.columns(2)
        b1.download_button("⬇️ Products (CSV)",     fe["products"].to_csv(index=False).encode(), "products.csv",        "text/csv")
        b2.download_button("⬇️ Audit trail (CSV)",  fe["detail"].to_csv(index=False).encode(),   "file_evidence_detail.csv", "text/csv")

# ── Software inventory path ────────────────────────────────────────────────
elif "Software inventory" in mode:
    if uploaded:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded.name)[1]) as tmp:
            tmp.write(uploaded.getvalue()); tmp_path = tmp.name
        try:
            raw_df = sniff_and_load(tmp_path)
        finally:
            os.unlink(tmp_path)

        schema = detect_schema(raw_df)
        st.success(f"Loaded {len(raw_df)} rows · {len(raw_df.columns)} columns")
        st.dataframe(raw_df.head(8), use_container_width=True)

        st.write("**Auto-detected columns** (adjust if wrong):")
        cols_none = ["(none)"] + list(raw_df.columns)
        def _pick(role, fallback=0):
            c = schema.get(role, {}).get("column")
            if c and c in raw_df.columns: return list(raw_df.columns).index(c) + 1
            return fallback
        c1, c2, c3 = st.columns(3)
        with c1: prod_col = st.selectbox("Product / Name column ★", cols_none, index=_pick("product", 1))
        with c2: vend_col = st.selectbox("Vendor / Publisher column", cols_none, index=_pick("vendor"))
        with c3: ver_col  = st.selectbox("Version column", cols_none, index=_pick("version"))
        c4, c5 = st.columns(2)
        with c4: host_col = st.selectbox("Host / Device column (optional)", cols_none, index=_pick("host"))
        with c5: ed_col   = st.selectbox("Edition column (optional)", cols_none, index=_pick("edition"))

        if st.button("Run normalisation 🚀", type="primary"):
            with st.spinner("Classifying…"):
                res = classify_inventory(
                    raw_df, norm,
                    name_col=prod_col if prod_col != "(none)" else raw_df.columns[0],
                    vendor_col=vend_col if vend_col != "(none)" else None,
                    version_col=ver_col  if ver_col  != "(none)" else None,
                    host_col=host_col   if host_col  != "(none)" else None,
                    edition_col=ed_col   if ed_col    != "(none)" else None,
                )
            st.session_state.results    = res
            st.session_state.llm_results = None

# ── Generic CSV path ──────────────────────────────────────────────────────
else:
    if uploaded:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded.name)[1]) as tmp:
            tmp.write(uploaded.getvalue()); tmp_path = tmp.name
        try:
            raw_df = sniff_and_load(tmp_path)
        finally:
            os.unlink(tmp_path)

        st.session_state.source_df = raw_df
        st.success(f"Loaded {len(raw_df)} rows · {len(raw_df.columns)} columns")
        st.dataframe(raw_df.head(8), use_container_width=True)

    if st.session_state.source_df is not None:
        raw_df = st.session_state.source_df
        st.subheader("2 · Map columns")
        prod_col = st.selectbox("Product name column ★", raw_df.columns)
        qty_col  = st.selectbox("Quantity column (optional)", ["(none)"] + list(raw_df.columns))

        if st.button("Run normalisation 🚀", type="primary"):
            with st.spinner("Classifying…"):
                res = norm.classify_batch(raw_df[prod_col].fillna("").astype(str))
                if qty_col != "(none)":
                    res.insert(0, "qty", pd.to_numeric(raw_df[qty_col], errors="coerce").fillna(0))
                    res["effective_qty"] = res["qty"] * res["suggested_qty_multiplier"]
            st.session_state.results    = res
            st.session_state.llm_results = None

# ── Results (shared by inventory + generic paths) ──────────────────────────
if st.session_state.results is not None and "File evidence" not in mode:
    results = st.session_state.results
    st.divider()
    st.subheader("Results")

    tier_counts = results["confidence_tier"].value_counts()
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total rows",       len(results))
    m2.metric("EXACT / HIGH",     int(tier_counts.get("EXACT",0) + tier_counts.get("HIGH",0)))
    m3.metric("REVIEW",           int(tier_counts.get("REVIEW",0)))
    m4.metric("PARSED (not in library)", int(tier_counts.get("PARSED",0)))
    m5.metric("UNRESOLVED",       int(tier_counts.get("UNRESOLVED",0)))

    # Qty multiplier warnings
    if "effective_qty" in results.columns:
        mult_rows = results[results["suggested_qty_multiplier"] > 1]
        if len(mult_rows):
            st.warning(
                f"⚠️ {len(mult_rows)} row(s) had an embedded count (e.g. '2 CPU', '2 Core Pack') "
                f"suggesting the reported quantity understates license units needed. "
                f"See `effective_qty` — this is a **naive multiplier, not a compliance calculation**. "
                f"Real per-core/socket minimums (Microsoft, Oracle, IBM, VMware) need human confirmation."
            )
            st.dataframe(
                mult_rows[["raw_input","qty","suggested_qty_multiplier","multiplier_source","effective_qty"]],
                use_container_width=True
            )

    # Filter & view
    tiers = ["EXACT","HIGH","REVIEW","PARSED","UNRESOLVED"]
    sel   = st.multiselect("Filter by confidence tier", tiers, default=tiers)
    filtered = results[results["confidence_tier"].isin(sel)] if sel else results
    st.write(f"Showing **{len(filtered)}** of {len(results)} rows")
    st.dataframe(filtered, use_container_width=True, height=400)

    # LLM tier for UNRESOLVED
    unresolved = results[results["confidence_tier"] == "UNRESOLVED"]
    if len(unresolved):
        st.subheader(f"🤖 Optional: LLM tier for {len(unresolved)} unresolved rows")
        st.caption("Uses Groq's free API to suggest matches — still needs human confirmation, never auto-accepted.")
        with st.expander("Configure Groq"):
            groq_key = st.text_input("Groq API key (or set GROQ_API_KEY env var)", type="password",
                                     help="Free key at https://console.groq.com")
            if st.button("Run LLM classification"):
                try:
                    from llm_classifier import classify_unresolved_batch
                    with st.spinner(f"Asking LLM about {len(unresolved)} rows…"):
                        llm_res = classify_unresolved_batch(
                            unresolved["raw_input"].tolist(), norm.lib,
                            api_key=groq_key or None
                        )
                    st.session_state.llm_results = llm_res
                    st.success(f"Suggestions received for {len(llm_res)} rows.")
                except Exception as e:
                    st.error(f"LLM call failed: {e}")

        if st.session_state.llm_results:
            llm = st.session_state.llm_results
            rows = []
            for raw, info in llm.items():
                cid = info.get("canonical_id")
                lib_row = norm.lib[norm.lib["canonical_id"] == cid]
                rows.append({
                    "raw_input": raw,
                    "llm_suggested": cid,
                    "llm_confidence": info.get("confidence"),
                    "llm_reasoning": info.get("reasoning"),
                    "suggested_product": (
                        f"{lib_row.iloc[0]['publisher']} | {lib_row.iloc[0]['product_family']}"
                        if len(lib_row) else "(not in library)"
                    )
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True)

    # Review queue
    review = results[results["confidence_tier"].isin(["REVIEW","UNRESOLVED"])]
    if len(review):
        st.subheader(f"✏️ Review & teach library ({len(review)} rows)")
        st.caption("Confirming a match permanently adds the alias — next engagement it resolves automatically.")
        canonical_opts = ["-- unresolved / not in library --"] + [
            f"{r.publisher} | {r.product_family} | {r.edition} ({r.canonical_id})"
            for r in norm.lib.itertuples()
        ]
        id_lookup = {opt: r.canonical_id for opt, r in zip(canonical_opts[1:], norm.lib.itertuples())}
        llm_res = st.session_state.llm_results or {}

        for idx, row in review.iterrows():
            llm_hint = llm_res.get(row["raw_input"])
            label = f"[{row['confidence_tier']}] {row['raw_input']}  (score: {row['match_score']})"
            if llm_hint and llm_hint.get("canonical_id"): label += "  🤖"
            with st.expander(label):
                if llm_hint and llm_hint.get("canonical_id"):
                    st.info(f"🤖 LLM: {llm_hint['canonical_id']} — {llm_hint.get('reasoning','')}")
                default_cid = (llm_hint or {}).get("canonical_id") or row["canonical_id"]
                default_idx = 0
                if default_cid:
                    for i, o in enumerate(canonical_opts):
                        if str(default_cid) in o: default_idx = i; break
                choice = st.selectbox("Correct match", canonical_opts, index=default_idx, key=f"sel_{idx}")
                if st.button("Confirm & teach library ✅", key=f"confirm_{idx}"):
                    if choice != "-- unresolved / not in library --":
                        norm.add_alias(id_lookup[choice], row["raw_input"])
                        norm.save(LIBRARY_PATH)
                        st.session_state.aliases_added_this_session = (
                            st.session_state.get("aliases_added_this_session", 0) + 1
                        )
                        st.success(
                            f"Saved. '{row['raw_input']}' → {id_lookup[choice]}  "
                            f"·  Remember to download the library from the sidebar "
                            f"and commit it to GitHub to keep this permanently."
                        )
                    else:
                        st.info("Left unresolved — add as new canonical product if it recurs.")

    # ── Export ────────────────────────────────────────────────────────
    st.subheader("Export")

    def build_excel(results: pd.DataFrame) -> bytes:
        """4-sheet workbook, consistent house style throughout.

           Overview      - run statistics (hidden by default)
           Normalised    - every row, host + product + classification, colour-coded
           Summary       - EVERY product grouped: publisher / product / type / count
           Product Count - data table + large horizontal bar chart
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

        # ── House style constants (used by every sheet) ────────────────
        NAVY   = "1F4E79"      # primary header
        GREEN  = "375623"      # summary header
        BLUE   = "2E75B6"      # chart / accent
        GREY   = "404040"      # overview header
        BAND   = "F2F2F2"      # zebra striping
        TIER_COLOURS = {
            "EXACT": "C6EFCE", "HIGH": "D9EAD3", "REVIEW": "FFF2CC",
            "PARSED": "DEEBF7", "UNRESOLVED": "F4CCCC",
        }
        CLASS_COLOURS = {
            "Commercial": "C6EFCE", "Component": "E4DFEC",
            "Freeware":   "DDEBF7", "Shareware": "FFF2CC",
        }
        thin = Side(style="thin", color="BFBFBF")
        BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(ws, colour):
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=colour)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 32

        def autosize(ws, cap=48):
            for col in ws.columns:
                width = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, cap)

        def finish(ws, freeze=True, autofilter=True):
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = BOX
            autosize(ws)
            if freeze:
                ws.freeze_panes = "A2"
            if autofilter and ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:

            # ══════════ SHEET 1: OVERVIEW (hidden) ════════════════════
            tiers = results["confidence_tier"].value_counts()
            n_rows   = len(results)
            n_exact  = int(tiers.get("EXACT", 0))
            n_high   = int(tiers.get("HIGH", 0))
            n_review = int(tiers.get("REVIEW", 0))
            n_parsed = int(tiers.get("PARSED", 0))
            n_unres  = int(tiers.get("UNRESOLVED", 0))
            n_hosts  = int(results["host_name"].nunique()) if "host_name" in results.columns else 0
            n_prod   = int(results[["publisher","product_family"]].drop_duplicates().shape[0])
            n_pub    = int(results["publisher"].nunique())

            cls = (results["classification"].fillna("Unclassified").value_counts()
                   if "classification" in results.columns else pd.Series(dtype=int))
            n_comm  = int(cls.get("Commercial", 0))
            n_compo = int(cls.get("Component", 0))
            n_free  = int(cls.get("Freeware", 0))
            n_share = int(cls.get("Shareware", 0))

            pct = lambda x: f"{x / n_rows * 100:.1f}%" if n_rows else "0.0%"

            ov_rows = [
                ("VOLUME", None, None),
                ("Total rows processed",           n_rows,  ""),
                ("Distinct hosts / devices",       n_hosts, ""),
                ("Distinct products",              n_prod,  ""),
                ("Distinct publishers",            n_pub,   ""),
                (None, None, None),
                ("MATCH CONFIDENCE", None, None),
                ("EXACT  - accept as-is",          n_exact,  pct(n_exact)),
                ("HIGH   - accept, spot-check",    n_high,   pct(n_high)),
                ("REVIEW - human must confirm",    n_review, pct(n_review)),
                ("PARSED - not in library",        n_parsed, pct(n_parsed)),
                ("UNRESOLVED - classify manually", n_unres,  pct(n_unres)),
                ("Auto-resolved (EXACT + HIGH)",   n_exact + n_high, pct(n_exact + n_high)),
                ("Needs human attention",          n_review + n_parsed + n_unres,
                                                   pct(n_review + n_parsed + n_unres)),
                (None, None, None),
                ("LICENSABILITY", None, None),
                ("Commercial - needs a licence",   n_comm,  pct(n_comm)),
                ("Component  - bundled, exclude",  n_compo, pct(n_compo)),
                ("Freeware   - free, exclude",     n_free,  pct(n_free)),
                ("Shareware  - trial / evaluation",n_share, pct(n_share)),
            ]
            pd.DataFrame(ov_rows, columns=["Metric", "Value", "% of Rows"]).to_excel(
                w, index=False, sheet_name="Overview")
            ws0 = w.sheets["Overview"]
            style_header(ws0, GREY)
            ws0.column_dimensions["A"].width = 40
            ws0.column_dimensions["B"].width = 14
            ws0.column_dimensions["C"].width = 14
            for row in ws0.iter_rows(min_row=2):
                label = str(row[0].value or "")
                if label in ("VOLUME", "MATCH CONFIDENCE", "LICENSABILITY"):
                    for c in row:
                        c.fill = PatternFill("solid", fgColor="D9D9D9")
                        c.font = Font(bold=True, size=11)
                elif label.startswith("Auto-resolved"):
                    for c in row:
                        c.fill = PatternFill("solid", fgColor="C6EFCE"); c.font = Font(bold=True)
                elif label.startswith("Needs human"):
                    for c in row:
                        c.fill = PatternFill("solid", fgColor="FFF2CC"); c.font = Font(bold=True)
                for c in row[1:]:
                    c.alignment = Alignment(horizontal="center")
            ws0.sheet_state = "hidden"

            # ══════════ SHEET 2: NORMALISED (host included) ═══════════
            s1_cols = [c for c in [
                "host_name", "raw_input", "publisher", "product_family",
                "edition", "version", "classification", "confidence_tier",
                "match_score", "metric_type", "source_vendor", "canonical_id",
            ] if c in results.columns]
            sheet1 = results[s1_cols].copy()
            sheet1.columns = [
                {"host_name": "Host", "raw_input": "Raw Product Name",
                 "publisher": "Publisher", "product_family": "Normalised Product",
                 "edition": "Edition", "version": "Version",
                 "classification": "Type", "confidence_tier": "Confidence",
                 "match_score": "Score", "metric_type": "Licence Metric",
                 "source_vendor": "Source Vendor", "canonical_id": "Canonical ID"}.get(c, c.title())
                for c in sheet1.columns
            ]
            sheet1.to_excel(w, index=False, sheet_name="Normalised")
            ws1 = w.sheets["Normalised"]
            style_header(ws1, NAVY)
            conf_idx = next((i for i, c in enumerate(ws1[1], 1)
                             if str(c.value) == "Confidence"), None)
            if conf_idx:
                for row in ws1.iter_rows(min_row=2):
                    colour = TIER_COLOURS.get(str(row[conf_idx - 1].value or ""))
                    if colour:
                        row[conf_idx - 1].fill = PatternFill("solid", fgColor=colour)
                        row[conf_idx - 1].font = Font(bold=True)
                        row[conf_idx - 1].alignment = Alignment(horizontal="center")
            # Zebra striping for readability across wide rows
            for i, row in enumerate(ws1.iter_rows(min_row=2), start=2):
                if i % 2 == 0:
                    for c in row:
                        if not c.fill or c.fill.fgColor.rgb in (None, "00000000"):
                            c.fill = PatternFill("solid", fgColor=BAND)
            finish(ws1)

            # ══════════ SHEET 3: SUMMARY (ALL products) ═══════════════
            # Every product, not just Commercial - the Type column lets the
            # reader filter to Commercial themselves, and seeing Components
            # and Freeware explicitly is what proves they were excluded on
            # purpose rather than silently lost.
            grp = [c for c in ["publisher", "product_family", "edition",
                                "version", "classification", "metric_type"]
                   if c in results.columns]
            agg = {"device_count": ("raw_input", "count")}
            if "host_name" in results.columns:
                agg["host_count"] = ("host_name", "nunique")
            agg["confidence"] = ("confidence_tier",
                                 lambda s: ", ".join(f"{k}:{v}" for k, v in s.value_counts().items()))

            summary = (results.groupby(grp, dropna=False)
                       .agg(**agg).reset_index())
            # Commercial first, then by count - most licence-relevant at the top
            cls_order = {"Commercial": 0, "Shareware": 1, "Freeware": 2, "Component": 3}
            if "classification" in summary.columns:
                summary["_o"] = summary["classification"].map(cls_order).fillna(9)
                summary = summary.sort_values(["_o", "device_count"],
                                              ascending=[True, False]).drop(columns="_o")
            else:
                summary = summary.sort_values("device_count", ascending=False)

            rename = {"publisher": "Publisher", "product_family": "Product",
                      "edition": "Edition", "version": "Version",
                      "classification": "Type", "metric_type": "Licence Metric",
                      "device_count": "Installs", "host_count": "Hosts",
                      "confidence": "Confidence"}
            summary.columns = [rename.get(c, c.title()) for c in summary.columns]
            summary.to_excel(w, index=False, sheet_name="Summary")
            ws2 = w.sheets["Summary"]
            style_header(ws2, GREEN)
            type_idx = next((i for i, c in enumerate(ws2[1], 1)
                             if str(c.value) == "Type"), None)
            if type_idx:
                for row in ws2.iter_rows(min_row=2):
                    colour = CLASS_COLOURS.get(str(row[type_idx - 1].value or ""))
                    if colour:
                        for c in row:
                            c.fill = PatternFill("solid", fgColor=colour)
                        row[type_idx - 1].font = Font(bold=True)
                        row[type_idx - 1].alignment = Alignment(horizontal="center")
            for row in ws2.iter_rows(min_row=2):
                for c in row:
                    if isinstance(c.value, (int, float)):
                        c.alignment = Alignment(horizontal="center")
            finish(ws2)

            # ══════════ SHEET 4: PRODUCT COUNT + CHART ════════════════
            cdf = (results.groupby(["publisher", "product_family", "classification"],
                                   dropna=False)
                   .agg(count=("raw_input", "count")).reset_index()
                   .sort_values("count", ascending=True).tail(20))
            if len(cdf):
                cdf["Product"] = (cdf["publisher"].fillna("").astype(str)
                                  + " – " + cdf["product_family"].fillna("").astype(str))
                chart_df = cdf[["Product", "classification", "count"]].rename(
                    columns={"classification": "Type", "count": "Installs"})
            else:
                chart_df = pd.DataFrame({"Product": ["No data"], "Type": [""], "Installs": [0]})

            chart_df.to_excel(w, index=False, sheet_name="Product Count")
            ws3 = w.sheets["Product Count"]
            style_header(ws3, BLUE)
            ws3.column_dimensions["A"].width = 48
            ws3.column_dimensions["B"].width = 14
            ws3.column_dimensions["C"].width = 11
            for row in ws3.iter_rows(min_row=2):
                colour = CLASS_COLOURS.get(str(row[1].value or ""))
                if colour:
                    for c in row:
                        c.fill = PatternFill("solid", fgColor=colour)
                row[2].alignment = Alignment(horizontal="center")
            finish(ws3, autofilter=False)

            n = len(chart_df)
            chart = BarChart()
            chart.type      = "bar"
            chart.grouping  = "clustered"
            chart.style     = 2
            chart.title     = "Installs by Product (top 20)"
            chart.x_axis.title = "Installs"
            chart.gapWidth  = 40
            data   = Reference(ws3, min_col=3, min_row=1, max_row=n + 1)
            labels = Reference(ws3, min_col=1, min_row=2, max_row=n + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            s = chart.series[0]
            s.graphicalProperties.solidFill     = BLUE
            s.graphicalProperties.line.solidFill = NAVY
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal  = True
            chart.dataLabels.position = "outEnd"
            chart.legend = None
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            # Size via anchor - chart.width/height do not survive the save cycle
            chart.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=4, colOff=0, row=1, rowOff=0),
                to=AnchorMarker(col=19, colOff=0, row=max(26, n * 2 + 4), rowOff=0))
            ws3.add_chart(chart)

        return buf.getvalue()

    csv_bytes  = results.to_csv(index=False).encode()
    excel_bytes = build_excel(results)

    b1, b2 = st.columns(2)
    b1.download_button("⬇️ CSV", csv_bytes, "normalized_results.csv", "text/csv")
    b2.download_button(
        "⬇️ Excel (3 sheets)", excel_bytes, "normalized_results.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(
        "**Sheet 1 – Normalised:** every row, colour-coded by confidence  ·  "
        "**Sheet 2 – Summary:** commercial products grouped by product with device count  ·  "
        "**Sheet 3 – Product Count:** horizontal bar chart of top 20 products"
    )
